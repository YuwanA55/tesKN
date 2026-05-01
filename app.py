import streamlit as st
import numpy as np
import pandas as pd
import math, io, pickle, warnings, base64
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

warnings.filterwarnings('ignore')

# ─── CONFIG ───────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Klasifikasi Kesulitan Soal",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── WARNA ────────────────────────────────────────────────────────────────────
BIRU_TUA="1F4E79"; BIRU_MID="2E75B6"; BIRU_MUDA="BDD7EE"
HIJAU_TUA="375623"; HIJAU_MID="70AD47"; HIJAU_MUDA="E2EFDA"
MERAH_MUDA="FCE4D6"; MERAH_TUA="9C0006"; KUNING="FFF2CC"
ABU="D9D9D9"; PUTIH="FFFFFF"; HITAM="000000"
C1_BG="C6EFCE"; C2_BG="FFEB9C"; C3_BG="FFC7CE"
C1_FG="375623"; C2_FG="9C6500"; C3_FG="9C0006"
KET=["Mudah","Sedang","Sulit"]
LABEL=["C1 (Mudah)","C2 (Sedang)","C3 (Sulit)"]
COLORS_K=['#2E75B6','#70AD47','#E00000']
MARKERS_K=['o','s','^']

# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    
    .main { background: #F0F4F8; }
    
    .hero-banner {
        background: linear-gradient(135deg, #1F4E79 0%, #2E75B6 50%, #70AD47 100%);
        border-radius: 16px; padding: 36px 40px; margin-bottom: 28px;
        color: white; text-align: center; box-shadow: 0 8px 32px rgba(31,78,121,0.3);
    }
    .hero-banner h1 { font-size: 2.2rem; font-weight: 700; margin: 0 0 8px 0; letter-spacing: -0.5px; }
    .hero-banner p  { font-size: 1.05rem; opacity: 0.9; margin: 0; }

    .nav-container {
        display: flex; gap: 8px; background: white;
        padding: 10px 14px; border-radius: 12px;
        box-shadow: 0 2px 12px rgba(0,0,0,0.08); margin-bottom: 24px;
        flex-wrap: wrap;
    }
    .nav-btn {
        padding: 9px 20px; border-radius: 8px; border: none; cursor: pointer;
        font-weight: 600; font-size: 0.88rem; transition: all 0.2s;
        background: #F0F4F8; color: #4A5568;
    }
    .nav-btn:hover { background: #BDD7EE; color: #1F4E79; }
    .nav-btn.active { background: #2E75B6; color: white; box-shadow: 0 3px 10px rgba(46,117,182,0.35); }

    .card {
        background: white; border-radius: 14px; padding: 24px 28px;
        box-shadow: 0 2px 16px rgba(0,0,0,0.07); margin-bottom: 20px;
    }
    .card-title {
        font-size: 1.1rem; font-weight: 700; color: #1F4E79;
        border-left: 4px solid #2E75B6; padding-left: 12px; margin-bottom: 16px;
    }
    
    .metric-row { display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 20px; }
    .metric-box {
        flex: 1; min-width: 130px; background: linear-gradient(135deg, #1F4E79, #2E75B6);
        border-radius: 12px; padding: 18px 16px; text-align: center; color: white;
    }
    .metric-box.green { background: linear-gradient(135deg, #375623, #70AD47); }
    .metric-box.orange { background: linear-gradient(135deg, #7F3F00, #ED7D31); }
    .metric-box.red { background: linear-gradient(135deg, #7B0000, #E00000); }
    .metric-box .val { font-size: 1.9rem; font-weight: 700; line-height: 1.1; }
    .metric-box .lbl { font-size: 0.78rem; opacity: 0.85; margin-top: 4px; }

    .badge-mudah  { background:#C6EFCE; color:#375623; padding:3px 10px; border-radius:99px; font-size:0.82rem; font-weight:600; }
    .badge-sedang { background:#FFEB9C; color:#9C6500; padding:3px 10px; border-radius:99px; font-size:0.82rem; font-weight:600; }
    .badge-sulit  { background:#FFC7CE; color:#9C0006; padding:3px 10px; border-radius:99px; font-size:0.82rem; font-weight:600; }

    .iter-header {
        background: #1F4E79; color: white; padding: 12px 18px;
        border-radius: 10px 10px 0 0; font-weight: 700; font-size: 1rem;
    }
    .iter-body { border: 1px solid #BDD7EE; border-top: none; border-radius: 0 0 10px 10px; padding: 16px; background: white; margin-bottom: 12px; }
    
    .upload-zone {
        border: 2px dashed #2E75B6; border-radius: 12px; padding: 40px;
        text-align: center; background: #F0F4F8; margin: 20px 0;
    }
    .upload-zone h3 { color: #2E75B6; margin: 0 0 8px 0; }
    .upload-zone p  { color: #718096; margin: 0; font-size: 0.9rem; }

    .result-pred {
        padding: 20px 24px; border-radius: 12px; text-align: center;
        font-size: 1.4rem; font-weight: 700; margin: 16px 0;
    }
    .pred-mudah { background: #C6EFCE; color: #375623; border: 2px solid #70AD47; }
    .pred-sedang { background: #FFEB9C; color: #9C6500; border: 2px solid #FFC000; }
    .pred-sulit  { background: #FFC7CE; color: #9C0006; border: 2px solid #E00000; }
    
    .stButton>button {
        background: linear-gradient(135deg, #2E75B6, #1F4E79);
        color: white; border: none; border-radius: 8px;
        font-weight: 600; padding: 10px 24px;
        box-shadow: 0 3px 10px rgba(46,117,182,0.3);
    }
    .stButton>button:hover { transform: translateY(-1px); box-shadow: 0 5px 16px rgba(46,117,182,0.4); }
    
    .download-btn {
        display: inline-block; padding: 10px 22px; border-radius: 8px;
        font-weight: 600; font-size: 0.9rem; text-decoration: none;
        margin: 6px; cursor: pointer;
    }
    .dl-excel { background: #198754; color: white; }
    .dl-model { background: #6f42c1; color: white; }
    
    div[data-testid="stSidebar"] { background: #1F4E79; }
    div[data-testid="stSidebar"] * { color: white !important; }
    div[data-testid="stSidebar"] .stSelectbox label { color: white !important; }
    
    footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ─── SESSION STATE ─────────────────────────────────────────────────────────────
if 'page' not in st.session_state: st.session_state.page = 'upload'
if 'hasil' not in st.session_state: st.session_state.hasil = None


# ══════════════════════════════════════════════════════════════════════════════
#  CORE ALGORITHM
# ══════════════════════════════════════════════════════════════════════════════

def load_input(file_obj):
    df_raw = pd.read_excel(file_obj, header=None)
    hrow = None
    for i, row in df_raw.iterrows():
        vals = [str(v).strip().lower() for v in row.values]
        if any(v in ['soal','no','no.','no soal'] for v in vals):
            hrow = i; break
    if hrow is None: raise ValueError("Kolom 'Soal' tidak ditemukan")
    df = pd.read_excel(file_obj, header=hrow)
    df.columns = df.columns.str.strip()
    rmap = {}
    for col in df.columns:
        lo = col.strip().lower()
        if lo in ['soal','no','no.','no soal']:   rmap[col] = 'Soal'
        elif 'persentase' in lo or lo in ['persen','%']: rmap[col] = 'Persentase'
        elif 'waktu' in lo:                        rmap[col] = 'Waktu'
    df = df.rename(columns=rmap)
    for col in ['Soal','Persentase','Waktu']:
        if col not in df.columns:
            raise ValueError(f"Kolom '{col}' tidak ditemukan dalam file Excel")
    df = df[['Soal','Persentase','Waktu']].dropna()
    df['Persentase'] = pd.to_numeric(df['Persentase'], errors='coerce')
    df['Waktu']      = pd.to_numeric(df['Waktu'],      errors='coerce')
    df = df.dropna().reset_index(drop=True)
    soal_list = []
    for i, v in enumerate(df['Soal'], 1):
        s = str(v).strip()
        soal_list.append(f"S{int(s)}" if s.isdigit() else
                         s.upper() if s.lower().startswith('s') and s[1:].isdigit()
                         else f"S{i}")
    df['Soal'] = soal_list
    df['Persentase'] = df['Persentase'].round(2)
    df['Waktu']      = df['Waktu'].round(2)
    return df

def euclidean(p1, p2):
    return math.sqrt((p1[0]-p2[0])**2 + (p1[1]-p2[1])**2)

def choose_centroids(data):
    df = data.copy()
    c1r = df.sort_values(['Persentase','Waktu'], ascending=[False,True]).iloc[0]
    c3r = df.sort_values(['Persentase','Waktu'], ascending=[True,False]).iloc[0]
    df['_dp'] = abs(df['Persentase'] - df['Persentase'].median())
    df['_dw'] = abs(df['Waktu']      - df['Waktu'].median())
    c2r = df.sort_values(['_dp','_dw']).iloc[0]
    return [(round(c1r['Persentase'],2), round(c1r['Waktu'],2)),
            (round(c2r['Persentase'],2), round(c2r['Waktu'],2)),
            (round(c3r['Persentase'],2), round(c3r['Waktu'],2))]

def run_kmeans(data):
    points    = list(zip(data['Persentase'], data['Waktu']))
    centroids = choose_centroids(data)
    history   = []
    for it in range(20):
        assignments, dists = [], []
        for p in points:
            d = [round(euclidean(p, centroids[k]), 4) for k in range(3)]
            assignments.append(d.index(min(d))); dists.append(d)
        new_cen = []
        for k in range(3):
            mems = [points[i] for i in range(len(points)) if assignments[i]==k]
            if mems:
                new_cen.append((round(sum(m[0] for m in mems)/len(mems),2),
                                round(sum(m[1] for m in mems)/len(mems),2)))
            else:
                new_cen.append(centroids[k])
        conv = (new_cen == centroids)
        history.append({'centroids':list(centroids),'assignments':assignments,
                        'distances':dists,'new_centroids':new_cen,'converged':conv})
        if conv: break
        centroids = new_cen
    return history

def stratified_split(data, test_ratio=0.25, random_state=42):
    np.random.seed(random_state)
    groups = defaultdict(list)
    for i, row in data.iterrows(): groups[row['Keterangan']].append(i)
    train_idx, test_idx, detail = [], [], {}
    for label in ['Mudah','Sedang','Sulit']:
        idxs = groups[label]; np.random.shuffle(idxs)
        n_test  = max(1, round(len(idxs)*test_ratio))
        n_train = len(idxs) - n_test
        train_idx += idxs[:n_train]; test_idx += idxs[n_train:]
        detail[label] = {'total':len(idxs),'train':n_train,'test':n_test}
    return (data.loc[train_idx].reset_index(drop=True),
            data.loc[test_idx].reset_index(drop=True), detail)

class GaussianNaiveBayes:
    def fit(self, X, y):
        self.classes_ = np.unique(y)
        self.priors_, self.means_, self.stds_ = {}, {}, {}
        for c in self.classes_:
            Xc = X[y==c]
            self.priors_[c] = len(Xc)/len(y)
            self.means_[c]  = Xc.mean(axis=0)
            self.stds_[c]   = Xc.std(axis=0)+1e-9

    def _gaussian(self, x, mean, std):
        return (1/(np.sqrt(2*np.pi)*std)) * np.exp(-((x-mean)**2)/(2*std**2))

    def _log_posterior(self, x):
        return {c: np.log(self.priors_[c]) +
                np.sum(np.log(self._gaussian(x, self.means_[c], self.stds_[c])))
                for c in self.classes_}

    def predict(self, X):
        return np.array([max(self._log_posterior(x), key=self._log_posterior(x).get)
                         for x in X])

    def predict_proba(self, X):
        hasil = []
        for x in X:
            lp  = self._log_posterior(x)
            arr = np.array([lp[c] for c in self.classes_], dtype=float)
            arr -= arr.max(); probs = np.exp(arr); probs /= probs.sum()
            hasil.append(probs)
        return np.array(hasil)

def cm_manual(y_true, y_pred, classes):
    cm = pd.DataFrame(0, index=classes, columns=classes)
    for t, p in zip(y_true, y_pred): cm.loc[t,p] += 1
    return cm

def metrics_manual(cm):
    hasil = {}
    for c in cm.index:
        tp=cm.loc[c,c]; fp=cm[c].sum()-tp; fn=cm.loc[c].sum()-tp
        pre = tp/(tp+fp) if (tp+fp)>0 else 0.0
        rec = tp/(tp+fn) if (tp+fn)>0 else 0.0
        f1  = (2*pre*rec)/(pre+rec) if (pre+rec)>0 else 0.0
        hasil[c] = [pre, rec, f1]
    return pd.DataFrame(hasil, index=["Precision","Recall","F1"]).T


# ══════════════════════════════════════════════════════════════════════════════
#  CHART FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def make_kmeans_scatter(it_num, it_data, points, soal_list, converged=False):
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.set_facecolor('#F8F9FA'); fig.patch.set_facecolor('white')
    cents=it_data['centroids']; assigns=it_data['assignments']
    for k in range(3):
        xs=[points[i][0] for i in range(len(points)) if assigns[i]==k]
        ys=[points[i][1] for i in range(len(points)) if assigns[i]==k]
        ax.scatter(xs, ys, c=COLORS_K[k], marker=MARKERS_K[k], s=120,
                   zorder=3, alpha=0.85, edgecolors='white', linewidths=0.8,
                   label=f'C{k+1} - {KET[k]}')
        for i in range(len(points)):
            if assigns[i]==k:
                ax.annotate(soal_list[i],(points[i][0],points[i][1]),
                            textcoords="offset points",xytext=(5,3),
                            fontsize=8,color=COLORS_K[k],fontweight='bold')
    for k in range(3):
        ax.scatter(cents[k][0],cents[k][1],c=COLORS_K[k],marker='*',
                   s=500,zorder=5,edgecolors='black',linewidths=1.2)
    sfx=' (KONVERGEN)' if converged else ''
    ax.set_xlabel('Persentase Jawaban Benar (%)',fontsize=11,fontweight='bold')
    ax.set_ylabel('Waktu Rata-rata (detik)',fontsize=11,fontweight='bold')
    ax.set_title(f'Iterasi {it_num} — K-Means Clustering{sfx}\nC1={cents[0]}  C2={cents[1]}  C3={cents[2]}',
                 fontsize=12,fontweight='bold',pad=12)
    patches=[mpatches.Patch(color=COLORS_K[k],label=f'C{k+1} - {KET[k]}') for k in range(3)]
    patches.append(mpatches.Patch(color='gray',label='★ Centroid'))
    ax.legend(handles=patches,loc='upper right',fontsize=9,framealpha=0.9)
    ax.grid(True,linestyle='--',alpha=0.4,color='#CCCCCC')
    pcts=[p[0] for p in points]; wkts=[p[1] for p in points]
    pad_x=(max(pcts)-min(pcts))*0.15+3; pad_y=(max(wkts)-min(wkts))*0.15+5
    ax.set_xlim(min(pcts)-pad_x, max(pcts)+pad_x+10)
    ax.set_ylim(min(wkts)-pad_y, max(wkts)+pad_y)
    plt.tight_layout(pad=1.5)
    return fig

def make_cm_heatmap(cm_df, classes):
    fig, ax = plt.subplots(figsize=(5,4))
    data_arr = cm_df.values.astype(float)
    im = ax.imshow(data_arr, cmap='Blues')
    ax.set_xticks(range(len(classes))); ax.set_yticks(range(len(classes)))
    ax.set_xticklabels(classes,fontsize=11,fontweight='bold')
    ax.set_yticklabels(classes,fontsize=11,fontweight='bold')
    ax.set_xlabel('Prediksi',fontsize=12,fontweight='bold')
    ax.set_ylabel('Aktual',  fontsize=12,fontweight='bold')
    ax.set_title('Confusion Matrix\nGaussian Naive Bayes',fontsize=13,fontweight='bold',pad=10)
    for i in range(len(classes)):
        for j in range(len(classes)):
            val=int(data_arr[i,j])
            color='white' if data_arr[i,j]>data_arr.max()*0.6 else 'black'
            ax.text(j,i,str(val),ha='center',va='center',fontsize=16,fontweight='bold',color=color)
    plt.colorbar(im, ax=ax); plt.tight_layout()
    return fig

def make_metrics_bar(met_df, acc):
    fig, ax = plt.subplots(figsize=(7,4.5))
    classes=met_df.index.tolist(); x=np.arange(len(classes)); w=0.25
    colors_bar=['#2E75B6','#70AD47','#ED7D31']
    for i,(col,clr) in enumerate(zip(['Precision','Recall','F1'],colors_bar)):
        vals=[met_df.loc[c,col] for c in classes]
        bars=ax.bar(x+i*w,vals,w,label=col,color=clr,alpha=0.85,edgecolor='white')
        for bar,val in zip(bars,vals):
            ax.text(bar.get_x()+bar.get_width()/2,bar.get_height()+0.01,
                    f'{val:.2f}',ha='center',va='bottom',fontsize=9,fontweight='bold')
    ax.set_xticks(x+w); ax.set_xticklabels(classes,fontsize=11,fontweight='bold')
    ax.set_ylim(0,1.15); ax.set_ylabel('Score',fontsize=11,fontweight='bold')
    ax.set_title(f'Precision / Recall / F1-Score per Kelas\nAkurasi = {acc*100:.2f}%',
                 fontsize=12,fontweight='bold',pad=10)
    ax.axhline(y=acc,color='red',linestyle='--',linewidth=1.5,label=f'Akurasi={acc*100:.2f}%')
    ax.legend(fontsize=10,framealpha=0.9); ax.grid(axis='y',linestyle='--',alpha=0.4)
    plt.tight_layout(); return fig

def make_scatter_nb(train, test, y_pred_test, classes):
    fig, ax = plt.subplots(figsize=(8,5.5))
    ax.set_facecolor('#F8F9FA'); fig.patch.set_facecolor('white')
    color_map={'Mudah':COLORS_K[0],'Sedang':COLORS_K[1],'Sulit':COLORS_K[2]}
    for c in classes:
        sub=train[train['Keterangan']==c]
        ax.scatter(sub['Persentase'],sub['Waktu'],c=color_map[c],
                   marker='o',s=80,alpha=0.6,edgecolors='white',linewidths=0.5)
    for (_,row),pred in zip(test.iterrows(),y_pred_test):
        c=row['Keterangan']; correct=(pred==c)
        ax.scatter(row['Persentase'],row['Waktu'],c=color_map[c],marker='D',s=120,
                   edgecolors='black' if correct else 'red',linewidths=2,zorder=5)
        ax.annotate(row['Soal'],(row['Persentase'],row['Waktu']),
                    textcoords="offset points",xytext=(5,3),fontsize=7,fontweight='bold')
    ax.set_xlabel('Persentase Jawaban Benar (%)',fontsize=11,fontweight='bold')
    ax.set_ylabel('Waktu Rata-rata (detik)',      fontsize=11,fontweight='bold')
    ax.set_title('Sebaran Data Train vs Test — Gaussian Naive Bayes\n'
                 'Lingkaran=Train  |  Berlian=Test  |  Border Merah=Salah',
                 fontsize=11,fontweight='bold',pad=10)
    patches=[mpatches.Patch(color=color_map[c],label=c) for c in classes]
    patches+=[mpatches.Patch(facecolor='white',edgecolor='black',label='Test Benar'),
              mpatches.Patch(facecolor='white',edgecolor='red',  label='Test Salah')]
    ax.legend(handles=patches,loc='upper right',fontsize=9,framealpha=0.9)
    ax.grid(True,linestyle='--',alpha=0.4,color='#CCCCCC')
    plt.tight_layout(); return fig

def make_prior_pie(model):
    fig, ax = plt.subplots(figsize=(5,4))
    labels=[f"{c}\n({v*100:.1f}%)" for c,v in model.priors_.items()]
    sizes=[v for v in model.priors_.values()]
    colors_pie=[COLORS_K[KET.index(c)] for c in model.priors_.keys()]
    ax.pie(sizes,labels=labels,colors=colors_pie,autopct='%1.1f%%',
           startangle=140,textprops={'fontsize':10,'fontweight':'bold'},
           wedgeprops={'edgecolor':'white','linewidth':2})
    ax.set_title('Distribusi Prior Probability\nper Kelas',fontsize=12,fontweight='bold')
    plt.tight_layout(); return fig


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def solid(h): return PatternFill("solid",start_color=h,fgColor=h)
def aln(h="center",wrap=False): return Alignment(horizontal=h,vertical="center",wrap_text=wrap)
def bdr():
    s=Side(style="thin"); return Border(left=s,right=s,top=s,bottom=s)

def sc(ws,r,c,v,bold=False,bg=None,fg=HITAM,h="center",wrap=False,sz=11,border=True):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(bold=bold,color=fg,size=sz,name="Calibri")
    if bg: cell.fill=solid(bg)
    cell.alignment=aln(h,wrap)
    if border: cell.border=bdr()
    return cell

def mg(ws,r1,c1,r2,c2,v,bold=False,bg=None,fg=HITAM,h="center",wrap=False,sz=11,italic=False):
    ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
    cell=ws.cell(row=r1,column=c1,value=v)
    cell.font=Font(bold=bold,color=fg,size=sz,italic=italic,name="Calibri")
    if bg: cell.fill=solid(bg)
    cell.alignment=aln(h,wrap); cell.border=bdr()
    return cell

def fig_to_xlimage(fig):
    buf=io.BytesIO(); fig.savefig(buf,format='png',dpi=130,bbox_inches='tight',
                                   facecolor='white'); plt.close(fig); buf.seek(0)
    return XLImage(buf)

def write_iterasi_sheet(ws, it_num, it_data, data, prev_assigns=None):
    soal=list(data['Soal']); pcts=list(data['Persentase']); wkts=list(data['Waktu'])
    cents=it_data['centroids']; assigns=it_data['assignments']
    dists=it_data['distances']; new_cen=it_data['new_centroids']; converged=it_data['converged']
    for col,w in zip('ABCDEFGHI',[14,14,14,14,14,12,12,12,16]):
        ws.column_dimensions[col].width=w
    row=1
    c_str=" | ".join(f"C{i+1}=({cents[i][0]},{cents[i][1]})" for i in range(3))
    mg(ws,row,1,row,9,f"ITERASI {it_num} — Centroid: {c_str}",bold=True,bg=BIRU_TUA,fg=PUTIH,sz=12)
    ws.row_dimensions[row].height=22; row+=2
    mg(ws,row,1,row,9,"CENTROID YANG DIGUNAKAN",bold=True,bg=BIRU_MUDA,fg=BIRU_TUA); row+=1
    for k in range(3):
        mg(ws,row,1,row,2,LABEL[k],bold=True,bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        mg(ws,row,3,row,5,f"Persentase = {cents[k][0]}%",bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        mg(ws,row,6,row,9,f"Waktu = {cents[k][1]} detik",bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k]); row+=1
    row+=1
    mg(ws,row,1,row,9,"d = SQRT((x1-y1)^2 + (x2-y2)^2)",italic=True,bg=KUNING,fg="7F7F7F"); row+=2
    for c,h in enumerate(["Soal","Persentase","Waktu","d(C1)","d(C2)","d(C3)","Cluster","Keterangan","Ubah?"],1):
        sc(ws,row,c,h,bold=True,bg=BIRU_TUA,fg=PUTIH)
    row+=1
    for i in range(len(soal)):
        k=assigns[i]; bg=[C1_BG,C2_BG,C3_BG][k]; fg=[C1_FG,C2_FG,C3_FG][k]
        changed=" ←" if prev_assigns and prev_assigns[i]!=k else ""
        sc(ws,row,1,soal[i]+changed,bg=bg,fg=fg); sc(ws,row,2,round(pcts[i],2),bg=bg,fg=fg)
        sc(ws,row,3,int(wkts[i]),bg=bg,fg=fg)
        sc(ws,row,4,dists[i][0],bg=bg,fg=fg); sc(ws,row,5,dists[i][1],bg=bg,fg=fg)
        sc(ws,row,6,dists[i][2],bg=bg,fg=fg); sc(ws,row,7,f"C{k+1}",bold=True,bg=bg,fg=fg)
        sc(ws,row,8,KET[k],bold=True,bg=bg,fg=fg)
        sc(ws,row,9,"Ya" if (prev_assigns and prev_assigns[i]!=k) else "",bg=bg,fg=fg); row+=1
    row+=1
    mg(ws,row,1,row,9,"RINGKASAN CLUSTER",bold=True,bg=BIRU_MUDA,fg=BIRU_TUA); row+=1
    for k in range(3):
        mems=[soal[i] for i in range(len(soal)) if assigns[i]==k]
        mg(ws,row,1,row,2,f"C{k+1}—{KET[k]}: {len(mems)} soal",bold=True,bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        mg(ws,row,3,row,9,", ".join(mems),bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k],h="left",wrap=True)
        ws.row_dimensions[row].height=18; row+=1
    row+=1
    mg(ws,row,1,row,9,"CENTROID BARU  (Ck = Σx / n)",bold=True,bg=BIRU_MUDA,fg=BIRU_TUA); row+=1
    for k in range(3):
        mp=[pcts[i] for i in range(len(soal)) if assigns[i]==k]
        mw=[wkts[i] for i in range(len(soal)) if assigns[i]==k]
        n=len(mp); sp=round(sum(mp),2); sw=round(sum(mw),2)
        bg=[C1_BG,C2_BG,C3_BG][k]; fg=[C1_FG,C2_FG,C3_FG][k]
        mg(ws,row,1,row,2,LABEL[k],bold=True,bg=bg,fg=fg)
        mg(ws,row,3,row,4,f"n={n}",bg=bg,fg=fg)
        mg(ws,row,5,row,6,f"Ck%={sp}/{n}={new_cen[k][0]}",bg=bg,fg=fg,wrap=True)
        mg(ws,row,7,row,9,f"CkW={int(sw)}/{n}={new_cen[k][1]}",bg=bg,fg=fg,wrap=True); row+=1
    row+=1
    if converged:
        mg(ws,row,1,row,9,"✓ KONVERGEN — Centroid tidak berubah. Proses selesai.",
           bold=True,bg=HIJAU_MUDA,fg=HIJAU_TUA)
    else:
        nc=" | ".join(f"C{i+1}=({new_cen[i][0]},{new_cen[i][1]})" for i in range(3))
        mg(ws,row,1,row,9,f"Centroid berubah → Lanjut iterasi: {nc}",bold=True,bg=MERAH_MUDA,fg=MERAH_TUA)

def generate_excel(data, history, train, test, split_detail, model,
                   y_pred, y_proba, y_test, cm_df, met_df, acc, classes):
    wb = Workbook()

    # Sheet: Hasil Clustering
    ws_cl = wb.active; ws_cl.title = "Hasil Clustering"
    for c,w in enumerate([10,14,14,12,12],1):
        ws_cl.column_dimensions[get_column_letter(c)].width=w
    mg(ws_cl,1,1,1,5,"HASIL K-MEANS CLUSTERING",bold=True,bg=BIRU_TUA,fg=PUTIH,sz=13)
    for ci,h in enumerate(["Soal","Persentase (%)","Waktu (detik)","Cluster","Keterangan"],1):
        sc(ws_cl,2,ci,h,bold=True,bg=BIRU_MID,fg=PUTIH)
    for ri,(_, row) in enumerate(data.iterrows(),3):
        k=KET.index(row['Keterangan']); bg=[C1_BG,C2_BG,C3_BG][k]; fg=[C1_FG,C2_FG,C3_FG][k]
        sc(ws_cl,ri,1,row['Soal'],bg=bg,fg=fg); sc(ws_cl,ri,2,row['Persentase'],bg=bg,fg=fg)
        sc(ws_cl,ri,3,int(row['Waktu']),bg=bg,fg=fg); sc(ws_cl,ri,4,row['Cluster'],bold=True,bg=bg,fg=fg)
        sc(ws_cl,ri,5,row['Keterangan'],bold=True,bg=bg,fg=fg)

    # Sheet per Iterasi
    prev=None
    for it_num, it_data in enumerate(history,1):
        sname=f"Iterasi {it_num}{'(Final)' if it_data['converged'] else ''}"
        ws_it=wb.create_sheet(sname); write_iterasi_sheet(ws_it,it_num,it_data,data,prev)
        prev=it_data['assignments'][:]

    # Sheet: Grafik K-Means
    ws_gk=wb.create_sheet("Grafik K-Means")
    ws_gk.sheet_view.showGridLines=False
    ws_gk.merge_cells('A1:N1'); c=ws_gk['A1']
    c.value='VISUALISASI K-MEANS'; c.font=Font(bold=True,size=14,color=PUTIH,name='Calibri')
    c.fill=solid(BIRU_TUA); c.alignment=aln()
    ws_gk.row_dimensions[1].height=30
    points=list(zip(data['Persentase'],data['Waktu'])); soal_list=list(data['Soal'])
    rp=2
    for it_num, it_data in enumerate(history,1):
        fig=make_kmeans_scatter(it_num,it_data,points,soal_list,it_data['converged'])
        buf=io.BytesIO(); fig.savefig(buf,format='png',dpi=130,bbox_inches='tight',facecolor='white')
        plt.close(fig); buf.seek(0)
        img=XLImage(buf); img.width=720; img.height=480
        ws_gk.add_image(img,f'A{rp}'); rp+=24

    # Sheet: Split
    ws_sp=wb.create_sheet("Stratified Split")
    for c,w in enumerate([10,12,14,14,14,16],1): ws_sp.column_dimensions[get_column_letter(c)].width=w
    row=1
    mg(ws_sp,row,1,row,6,"STRATIFIED SPLIT 75:25",bold=True,bg=BIRU_TUA,fg=PUTIH,sz=13); row+=1
    mg(ws_sp,row,1,row,6,f"Total={len(data)}  Train={len(train)}(75%)  Test={len(test)}(25%)",
       bold=True,bg=BIRU_MUDA,fg=BIRU_TUA); row+=2
    for ci,h in enumerate(["Kelas","Total","Train","Test","% Train","% Test"],1):
        sc(ws_sp,row,ci,h,bold=True,bg=BIRU_TUA,fg=PUTIH)
    row+=1
    for label in ['Mudah','Sedang','Sulit']:
        d=split_detail[label]; k=KET.index(label)
        sc(ws_sp,row,1,label,bold=True,bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_sp,row,2,d['total'],bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_sp,row,3,d['train'],bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_sp,row,4,d['test'],bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_sp,row,5,f"{d['train']/d['total']*100:.1f}%",bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_sp,row,6,f"{d['test']/d['total']*100:.1f}%",bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k]); row+=1

    # Sheet: Training
    ws_tr=wb.create_sheet("Training NB")
    for c,w in enumerate([16,18,18,18,16],1): ws_tr.column_dimensions[get_column_letter(c)].width=w
    row=1
    mg(ws_tr,row,1,row,5,"TRAINING — GAUSSIAN NAIVE BAYES",bold=True,bg=BIRU_TUA,fg=PUTIH,sz=13); row+=2
    mg(ws_tr,row,1,row,5,"PRIOR PROBABILITY",bold=True,bg=BIRU_MID,fg=PUTIH); row+=1
    for ci,h in enumerate(["Kelas","Jumlah","Total","P(Ck)"],1):
        sc(ws_tr,row,ci,h,bold=True,bg=BIRU_TUA,fg=PUTIH)
    row+=1; total=len(train)
    for c in classes:
        k=KET.index(c); n=list(train['Keterangan']).count(c)
        sc(ws_tr,row,1,c,bold=True,bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_tr,row,2,n,bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_tr,row,3,total,bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_tr,row,4,round(model.priors_[c],4),bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k]); row+=1
    row+=1
    mg(ws_tr,row,1,row,5,"MEAN & STD PER KELAS",bold=True,bg=BIRU_MID,fg=PUTIH); row+=1
    for ci,h in enumerate(["Kelas","Mean %","Mean Waktu","Std %","Std Waktu"],1):
        sc(ws_tr,row,ci,h,bold=True,bg=BIRU_TUA,fg=PUTIH)
    row+=1
    for c in classes:
        k=KET.index(c)
        sc(ws_tr,row,1,c,bold=True,bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_tr,row,2,round(model.means_[c][0],4),bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_tr,row,3,round(model.means_[c][1],4),bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_tr,row,4,round(model.stds_[c][0],4),bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_tr,row,5,round(model.stds_[c][1],4),bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k]); row+=1

    # Sheet: Testing
    ws_te=wb.create_sheet("Testing NB")
    for c,w in enumerate([8,10,14,14,14,16,16,16,14,12],1): ws_te.column_dimensions[get_column_letter(c)].width=w
    row=1
    mg(ws_te,row,1,row,10,"HASIL PREDIKSI DATA UJI",bold=True,bg=BIRU_TUA,fg=PUTIH,sz=13); row+=2
    hdrs=["No","Soal","Persentase","Waktu","Aktual",
          f"P({classes[0]})",f"P({classes[1]})",f"P({classes[2]})","Prediksi","Status"]
    for ci,h in enumerate(hdrs,1): sc(ws_te,row,ci,h,bold=True,bg=BIRU_TUA,fg=PUTIH)
    row+=1
    for i,((_,r),pred,prob) in enumerate(zip(test.iterrows(),y_pred,y_proba),1):
        actual=r['Keterangan']; correct=(pred==actual)
        k_pred=KET.index(pred); bg=[C1_BG,C2_BG,C3_BG][k_pred]; fg=[C1_FG,C2_FG,C3_FG][k_pred]
        sbg=HIJAU_MUDA if correct else MERAH_MUDA; sfg=HIJAU_TUA if correct else MERAH_TUA
        sc(ws_te,row,1,i,bg=bg,fg=fg); sc(ws_te,row,2,r['Soal'],bg=bg,fg=fg)
        sc(ws_te,row,3,r['Persentase'],bg=bg,fg=fg); sc(ws_te,row,4,int(r['Waktu']),bg=bg,fg=fg)
        sc(ws_te,row,5,actual,bold=True,bg=bg,fg=fg)
        for ci2,p in enumerate(prob,6): sc(ws_te,row,ci2,f"{p*100:.2f}%",bg=bg,fg=fg)
        sc(ws_te,row,9,pred,bold=True,bg=bg,fg=fg)
        sc(ws_te,row,10,"BENAR" if correct else "SALAH",bold=True,bg=sbg,fg=sfg); row+=1

    # Sheet: Evaluasi
    ws_ev=wb.create_sheet("Evaluasi NB")
    for c,w in enumerate([16,14,14,14,14,14],1): ws_ev.column_dimensions[get_column_letter(c)].width=w
    row=1
    mg(ws_ev,row,1,row,6,"EVALUASI MODEL — GAUSSIAN NAIVE BAYES",bold=True,bg=BIRU_TUA,fg=PUTIH,sz=13); row+=1
    mg(ws_ev,row,1,row,6,f"Akurasi: {acc*100:.2f}%",bold=True,bg=HIJAU_MUDA,fg=HIJAU_TUA,sz=12); row+=2
    mg(ws_ev,row,1,row,6,"CONFUSION MATRIX",bold=True,bg=BIRU_MID,fg=PUTIH); row+=1
    sc(ws_ev,row,1,"Aktual\\Prediksi",bold=True,bg=BIRU_TUA,fg=PUTIH)
    for ci2,c in enumerate(classes,2): sc(ws_ev,row,ci2,c,bold=True,bg=BIRU_TUA,fg=PUTIH)
    row+=1
    for ra in classes:
        k=KET.index(ra); sc(ws_ev,row,1,ra,bold=True,bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        for ci2,ca in enumerate(classes,2):
            val=int(cm_df.loc[ra,ca])
            sc(ws_ev,row,ci2,val,bold=(ra==ca),bg=HIJAU_MUDA if ra==ca else MERAH_MUDA,
               fg=HIJAU_TUA if ra==ca else MERAH_TUA); row+=1
    row+=1
    mg(ws_ev,row,1,row,6,"PRECISION / RECALL / F1",bold=True,bg=BIRU_MID,fg=PUTIH); row+=1
    for ci2,h in enumerate(["Kelas","Precision","Recall","F1","Support"],1):
        sc(ws_ev,row,ci2,h,bold=True,bg=BIRU_TUA,fg=PUTIH)
    row+=1
    for c in classes:
        k=KET.index(c); sup=list(y_test).count(c)
        sc(ws_ev,row,1,c,bold=True,bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_ev,row,2,f"{met_df.loc[c,'Precision']:.4f}",bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_ev,row,3,f"{met_df.loc[c,'Recall']:.4f}",bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_ev,row,4,f"{met_df.loc[c,'F1']:.4f}",bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k])
        sc(ws_ev,row,5,sup,bg=[C1_BG,C2_BG,C3_BG][k],fg=[C1_FG,C2_FG,C3_FG][k]); row+=1
    row+=2
    mg(ws_ev,row,1,row,6,"GRAFIK EVALUASI",bold=True,bg=BIRU_MID,fg=PUTIH); row+=1
    fig_cm=make_cm_heatmap(cm_df,classes); buf=io.BytesIO()
    fig_cm.savefig(buf,format='png',dpi=130,bbox_inches='tight',facecolor='white'); plt.close(fig_cm); buf.seek(0)
    img_cm=XLImage(buf); img_cm.width=380; img_cm.height=310; ws_ev.add_image(img_cm,f'A{row}')
    fig_bar=make_metrics_bar(met_df,acc); buf2=io.BytesIO()
    fig_bar.savefig(buf2,format='png',dpi=130,bbox_inches='tight',facecolor='white'); plt.close(fig_bar); buf2.seek(0)
    img_bar=XLImage(buf2); img_bar.width=510; img_bar.height=330; ws_ev.add_image(img_bar,f'G{row}')

    xl_buf=io.BytesIO(); wb.save(xl_buf); xl_buf.seek(0)
    return xl_buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  RUN PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

def run_pipeline(file_obj):
    data    = load_input(file_obj)
    history = run_kmeans(data)
    final   = history[-1]
    data['Cluster']    = [f"C{a+1}" for a in final['assignments']]
    data['Keterangan'] = [KET[a]    for a in final['assignments']]
    train, test, split_detail = stratified_split(data, test_ratio=0.25)
    classes = sorted(data['Keterangan'].unique(), key=lambda x: KET.index(x))
    X_train=train[['Persentase','Waktu']].values; y_train=train['Keterangan'].values
    model=GaussianNaiveBayes(); model.fit(X_train, y_train)
    X_test=test[['Persentase','Waktu']].values;  y_test=test['Keterangan'].values
    y_pred=model.predict(X_test); y_proba=model.predict_proba(X_test)
    acc   =np.mean(y_test==y_pred)
    cm_df =cm_manual(y_test,y_pred,classes)
    met_df=metrics_manual(cm_df)
    model_buf=io.BytesIO(); pickle.dump(model,model_buf); model_buf.seek(0)
    xl_bytes=generate_excel(data,history,train,test,split_detail,model,
                            y_pred,y_proba,y_test,cm_df,met_df,acc,classes)
    return {'data':data,'history':history,'train':train,'test':test,
            'split_detail':split_detail,'model':model,'y_pred':y_pred,
            'y_proba':y_proba,'y_test':y_test,'cm_df':cm_df,'met_df':met_df,
            'acc':acc,'classes':classes,'model_buf':model_buf.getvalue(),
            'xl_bytes':xl_bytes}


# ══════════════════════════════════════════════════════════════════════════════
#  UI HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def badge(label):
    cls = {'Mudah':'badge-mudah','Sedang':'badge-sedang','Sulit':'badge-sulit'}.get(label,'badge-sedang')
    return f'<span class="{cls}">{label}</span>'

def b64dl(data_bytes, filename, label, css_class):
    b64=base64.b64encode(data_bytes).decode()
    ext=filename.split('.')[-1]
    mime={'xlsx':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
          'pkl':'application/octet-stream'}.get(ext,'application/octet-stream')
    return (f'<a href="data:{mime};base64,{b64}" download="{filename}" '
            f'class="download-btn {css_class}">{label}</a>')


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE RENDERERS
# ══════════════════════════════════════════════════════════════════════════════

def page_upload():
    st.markdown("""
    <div class="hero-banner">
        <h1>📊 Klasifikasi Tingkat Kesulitan Soal</h1>
        <p>Pipeline K-Means Clustering → Gaussian Naive Bayes</p>
    </div>""", unsafe_allow_html=True)

    col1, col2 = st.columns([3,2])
    with col1:
        st.markdown('<div class="card"><div class="card-title">📂 Upload File Excel</div>', unsafe_allow_html=True)
        uploaded = st.file_uploader("Pilih file Excel (.xlsx / .xls)",
                                    type=['xlsx','xls'], label_visibility='collapsed')
        if uploaded:
            st.success(f"✅ File berhasil dimuat: **{uploaded.name}**")
            try:
                preview = pd.read_excel(uploaded, header=None)
                st.dataframe(preview.head(8), use_container_width=True)
                uploaded.seek(0)
            except: pass
            if st.button("🚀 Jalankan Analisis", use_container_width=True):
                with st.spinner("Memproses data... Mohon tunggu"):
                    try:
                        uploaded.seek(0)
                        hasil = run_pipeline(uploaded)
                        st.session_state.hasil = hasil
                        st.session_state.page  = 'kmeans'
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {e}")
        else:
            st.markdown("""
            <div class="upload-zone">
                <h3>📤 Drag & Drop atau Klik untuk Upload</h3>
                <p>Format: Excel (.xlsx / .xls)<br>
                Kolom yang diperlukan: <strong>Soal, Persentase, Waktu</strong></p>
            </div>""", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with col2:
        st.markdown("""
        <div class="card">
        <div class="card-title">📋 Format Data</div>
        <p style="color:#4A5568; font-size:0.9rem;">File Excel harus memiliki kolom:</p>
        """, unsafe_allow_html=True)
        sample = pd.DataFrame({'Soal':['S1','S2','S3'],
                               'Persentase':[85.5,60.0,35.2],
                               'Waktu':[45,80,120]})
        st.dataframe(sample, use_container_width=True, hide_index=True)
        st.markdown("""
        <ul style="color:#4A5568; font-size:0.87rem; line-height:1.9;">
          <li><b>Soal</b> — Nomor/kode soal</li>
          <li><b>Persentase</b> — % jawaban benar siswa</li>
          <li><b>Waktu</b> — Waktu rata-rata pengerjaan (detik)</li>
        </ul>
        <hr style="border-color:#E2E8F0;">
        <p style="color:#718096; font-size:0.85rem;">
          <b>Algoritma:</b><br>
          1️⃣ K-Means (3 cluster: Mudah/Sedang/Sulit)<br>
          2️⃣ Gaussian Naive Bayes (train 75% / test 25%)<br>
          3️⃣ Evaluasi: Akurasi, Precision, Recall, F1
        </p>
        </div>""", unsafe_allow_html=True)


def page_kmeans():
    h = st.session_state.hasil
    data=h['data']; history=h['history']; final=history[-1]
    
    st.markdown('<div class="card"><div class="card-title">📍 Ringkasan K-Means Clustering</div>', unsafe_allow_html=True)
    counts={KET[k]:final['assignments'].count(k) for k in range(3)}
    c1,c2,c3,c4=st.columns(4)
    with c1: st.markdown(f'<div class="metric-box"><div class="val">{len(data)}</div><div class="lbl">Total Soal</div></div>',unsafe_allow_html=True)
    with c2: st.markdown(f'<div class="metric-box green"><div class="val">{counts["Mudah"]}</div><div class="lbl">C1 — Mudah</div></div>',unsafe_allow_html=True)
    with c3: st.markdown(f'<div class="metric-box orange"><div class="val">{counts["Sedang"]}</div><div class="lbl">C2 — Sedang</div></div>',unsafe_allow_html=True)
    with c4: st.markdown(f'<div class="metric-box red"><div class="val">{counts["Sulit"]}</div><div class="lbl">C3 — Sulit</div></div>',unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # Tabel hasil
    st.markdown('<div class="card"><div class="card-title">📋 Data Hasil Clustering</div>', unsafe_allow_html=True)
    df_show = data.copy()
    st.dataframe(df_show[['Soal','Persentase','Waktu','Cluster','Keterangan']],
                 use_container_width=True, hide_index=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # Iterasi detail
    st.markdown('<div class="card"><div class="card-title">🔄 Detail Iterasi K-Means</div>', unsafe_allow_html=True)
    tab_labels=[f"Iterasi {i+1}{'✅' if h['converged'] else ''}" for i,h in enumerate(history)]
    tabs=st.tabs(tab_labels)
    points=list(zip(data['Persentase'],data['Waktu'])); soal_list=list(data['Soal'])
    for ti,(tab,it_data) in enumerate(zip(tabs,history)):
        with tab:
            cents=it_data['centroids']; assigns=it_data['assignments']; dists=it_data['distances']
            cc1,cc2=st.columns([1,2])
            with cc1:
                st.markdown("**Centroid yang digunakan:**")
                for k in range(3):
                    clr=['🟢','🟡','🔴'][k]
                    st.markdown(f"{clr} **{LABEL[k]}**: P={cents[k][0]}%, W={cents[k][1]}s")
                st.markdown("---")
                cnts=[assigns.count(k) for k in range(3)]
                for k in range(3):
                    st.markdown(f"**{KET[k]}**: {cnts[k]} soal")
                nc=it_data['new_centroids']
                if it_data['converged']:
                    st.success("✅ KONVERGEN!")
                else:
                    st.info(f"Centroid baru → lanjut iterasi")
            with cc2:
                fig=make_kmeans_scatter(ti+1,it_data,points,soal_list,it_data['converged'])
                st.pyplot(fig,use_container_width=True); plt.close()
            
            st.markdown("**Tabel Jarak per Soal:**")
            rows=[]
            for i in range(len(soal_list)):
                k=assigns[i]
                rows.append({'Soal':soal_list[i],'Persentase':data['Persentase'].iloc[i],
                             'Waktu':int(data['Waktu'].iloc[i]),
                             'd(C1)':dists[i][0],'d(C2)':dists[i][1],'d(C3)':dists[i][2],
                             'Cluster':f"C{k+1}",'Keterangan':KET[k]})
            st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
    st.markdown('</div>', unsafe_allow_html=True)


def page_model():
    h = st.session_state.hasil
    model=h['model']; train=h['train']; test=h['test']; classes=h['classes']
    split_detail=h['split_detail']

    st.markdown('<div class="card"><div class="card-title">✂️ Stratified Split 75% : 25%</div>', unsafe_allow_html=True)
    c1,c2,c3=st.columns(3)
    with c1: st.markdown(f'<div class="metric-box"><div class="val">{len(train)}</div><div class="lbl">Data Train (75%)</div></div>',unsafe_allow_html=True)
    with c2: st.markdown(f'<div class="metric-box green"><div class="val">{len(test)}</div><div class="lbl">Data Test (25%)</div></div>',unsafe_allow_html=True)
    with c3: st.markdown(f'<div class="metric-box orange"><div class="val">{len(train)+len(test)}</div><div class="lbl">Total Data</div></div>',unsafe_allow_html=True)
    st.markdown('</br>',unsafe_allow_html=True)
    rows=[{'Kelas':k,'Total':split_detail[k]['total'],'Train':split_detail[k]['train'],
           'Test':split_detail[k]['test'],
           '% Train':f"{split_detail[k]['train']/split_detail[k]['total']*100:.1f}%",
           '% Test':f"{split_detail[k]['test']/split_detail[k]['total']*100:.1f}%"}
          for k in ['Mudah','Sedang','Sulit']]
    st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="card"><div class="card-title">🧠 Parameter Model Gaussian Naive Bayes</div>', unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        st.markdown("**Prior Probability P(Ck):**")
        prior_rows=[{'Kelas':c,'Jumlah Train':list(train['Keterangan']).count(c),
                     'Total':len(train),'P(Ck)':f"{model.priors_[c]:.4f}"}
                    for c in classes]
        st.dataframe(pd.DataFrame(prior_rows),use_container_width=True,hide_index=True)
        fig_pie=make_prior_pie(model); st.pyplot(fig_pie,use_container_width=True); plt.close()
    with c2:
        st.markdown("**Mean & Std per Kelas:**")
        ms_rows=[{'Kelas':c,
                  'Mean Persentase':round(model.means_[c][0],4),
                  'Mean Waktu':round(model.means_[c][1],4),
                  'Std Persentase':round(model.stds_[c][0],4),
                  'Std Waktu':round(model.stds_[c][1],4)}
                 for c in classes]
        st.dataframe(pd.DataFrame(ms_rows),use_container_width=True,hide_index=True)
        st.markdown("""
        <div style="background:#EFF6FF;border-radius:10px;padding:14px;margin-top:12px;">
        <b>📐 Rumus Gaussian PDF:</b><br>
        <code style="font-size:0.85rem;">P(x|Ck) = (1/√(2π)σ) × exp(-(x-μ)² / 2σ²)</code>
        <br><br><b>📝 Log Posterior:</b><br>
        <code style="font-size:0.85rem;">log P(Ck|x) = log P(Ck) + Σ log P(xi|Ck)</code>
        <br><br><b>🎯 Prediksi:</b><br>
        <code style="font-size:0.85rem;">Kelas = argmax [ log P(Ck|x) ]</code>
        </div>""", unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # Tabel prediksi test
    st.markdown('<div class="card"><div class="card-title">🔮 Hasil Prediksi Data Test</div>', unsafe_allow_html=True)
    y_pred=h['y_pred']; y_proba=h['y_proba']
    rows=[]
    for i,((_,r),pred,prob) in enumerate(zip(test.iterrows(),y_pred,y_proba),1):
        rows.append({'No':i,'Soal':r['Soal'],'Persentase':r['Persentase'],
                     'Waktu':int(r['Waktu']),'Aktual':r['Keterangan'],
                     f'P({classes[0]})':f"{prob[0]*100:.2f}%",
                     f'P({classes[1]})':f"{prob[1]*100:.2f}%",
                     f'P({classes[2]})':f"{prob[2]*100:.2f}%",
                     'Prediksi':pred,'Status':'✅ Benar' if pred==r['Keterangan'] else '❌ Salah'})
    st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
    st.markdown('</div>', unsafe_allow_html=True)


def page_evaluasi():
    h = st.session_state.hasil
    acc=h['acc']; cm_df=h['cm_df']; met_df=h['met_df']
    classes=h['classes']; y_test=h['y_test']; y_pred=h['y_pred']
    train=h['train']; test=h['test']

    acc_pct=acc*100
    acc_color="green" if acc_pct>=80 else ("orange" if acc_pct>=60 else "red")

    st.markdown(f"""
    <div class="card" style="text-align:center;">
    <div class="card-title" style="text-align:left;">🎯 Akurasi Model</div>
    <div style="font-size:4rem;font-weight:800;color:{'#375623' if acc_color=='green' else '#9C6500' if acc_color=='orange' else '#9C0006'};">
        {acc_pct:.2f}%
    </div>
    <div style="font-size:1rem;color:#718096;margin-top:6px;">
        {sum(1 for t,p in zip(y_test,y_pred) if t==p)} prediksi benar dari {len(y_test)} data uji
    </div>
    </div>""", unsafe_allow_html=True)

    c1,c2,c3=st.columns(3)
    with c1: st.markdown(f'<div class="metric-box"><div class="val">{met_df["Precision"].mean():.4f}</div><div class="lbl">Macro Precision</div></div>',unsafe_allow_html=True)
    with c2: st.markdown(f'<div class="metric-box green"><div class="val">{met_df["Recall"].mean():.4f}</div><div class="lbl">Macro Recall</div></div>',unsafe_allow_html=True)
    with c3: st.markdown(f'<div class="metric-box orange"><div class="val">{met_df["F1"].mean():.4f}</div><div class="lbl">Macro F1-Score</div></div>',unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    # Metrics per kelas
    st.markdown('<div class="card"><div class="card-title">📊 Metrik per Kelas</div>', unsafe_allow_html=True)
    met_show=[{'Kelas':c,'Precision':f"{met_df.loc[c,'Precision']:.4f}",
               'Recall':f"{met_df.loc[c,'Recall']:.4f}",'F1-Score':f"{met_df.loc[c,'F1']:.4f}",
               'Support':list(y_test).count(c)} for c in classes]
    st.dataframe(pd.DataFrame(met_show),use_container_width=True,hide_index=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # Visualisasi
    st.markdown('<div class="card"><div class="card-title">📈 Visualisasi Evaluasi</div>', unsafe_allow_html=True)
    vc1,vc2=st.columns(2)
    with vc1:
        st.markdown("**Confusion Matrix**")
        fig_cm=make_cm_heatmap(cm_df,classes); st.pyplot(fig_cm,use_container_width=True); plt.close()
    with vc2:
        st.markdown("**Precision / Recall / F1 per Kelas**")
        fig_bar=make_metrics_bar(met_df,acc); st.pyplot(fig_bar,use_container_width=True); plt.close()
    st.markdown("**Sebaran Data Train vs Test**")
    fig_sc=make_scatter_nb(train,test,y_pred,classes); st.pyplot(fig_sc,use_container_width=True); plt.close()
    st.markdown('</div>', unsafe_allow_html=True)


def page_prediksi():
    h = st.session_state.hasil
    model=h['model']; classes=h['classes']

    st.markdown('<div class="card"><div class="card-title">🔮 Prediksi Tingkat Kesulitan Soal Baru</div>', unsafe_allow_html=True)
    st.markdown("Masukkan data soal yang ingin diklasifikasikan:")
    
    c1,c2,c3=st.columns(3)
    with c1:
        total_siswa=st.number_input("👥 Jumlah Siswa Total",min_value=1,value=30,step=1)
    with c2:
        jumlah_benar=st.number_input("✅ Jumlah Jawaban Benar",min_value=0,value=20,step=1)
    with c3:
        waktu=st.number_input("⏱️ Waktu Rata-rata (detik)",min_value=1,value=60,step=1)

    if jumlah_benar > total_siswa:
        st.error("⚠️ Jumlah jawaban benar tidak boleh melebihi jumlah siswa!")
    else:
        persentase = round((jumlah_benar / total_siswa) * 100, 2)
        st.info(f"📐 Persentase dihitung otomatis: **{jumlah_benar}/{total_siswa} × 100 = {persentase}%**")

        if st.button("🎯 Prediksi Sekarang", use_container_width=True):
            x = np.array([[persentase, waktu]])
            pred = model.predict(x)[0]
            proba = model.predict_proba(x)[0]

            cls_idx = pred
            pred_css = {'Mudah':'pred-mudah','Sedang':'pred-sedang','Sulit':'pred-sulit'}.get(pred,'pred-sedang')
            emoji = {'Mudah':'🟢','Sedang':'🟡','Sulit':'🔴'}.get(pred,'⚪')

            st.markdown(f"""
            <div class="result-pred {pred_css}">
                {emoji} Soal ini diprediksi: <strong>{pred}</strong>
            </div>""", unsafe_allow_html=True)

            st.markdown("**Probabilitas per kelas:**")
            for c, p in zip(model.classes_, proba):
                emoji_c={'Mudah':'🟢','Sedang':'🟡','Sulit':'🔴'}.get(c,'⚪')
                bar_color={'Mudah':'#70AD47','Sedang':'#FFC000','Sulit':'#E00000'}.get(c,'#2E75B6')
                st.markdown(f"{emoji_c} **{c}**: {p*100:.2f}%")
                st.progress(float(p))

            st.markdown(f"""
            <div style="background:#F0F4F8;border-radius:10px;padding:14px;margin-top:12px;">
            <b>📋 Ringkasan Input:</b><br>
            • Jumlah Siswa: {total_siswa} | Jawaban Benar: {jumlah_benar} | Persentase: {persentase}% | Waktu: {waktu} detik
            </div>""", unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    # Batch prediction
    st.markdown('<div class="card"><div class="card-title">📦 Prediksi Batch (Multi Soal)</div>', unsafe_allow_html=True)
    st.markdown("Masukkan beberapa soal sekaligus (format: `NomorSoal, TotalSiswa, Benar, Waktu`):")
    batch_input = st.text_area("Input Batch",
        placeholder="S1, 30, 25, 45\nS2, 30, 12, 90\nS3, 30, 5, 120",
        height=120)
    if st.button("🔮 Prediksi Batch"):
        try:
            rows=[]
            for line in batch_input.strip().split('\n'):
                parts=[p.strip() for p in line.split(',')]
                if len(parts)==4:
                    soal,tot,btr,wkt=parts[0],int(parts[1]),int(parts[2]),float(parts[3])
                    pct=round(btr/tot*100,2)
                    pred=model.predict(np.array([[pct,wkt]]))[0]
                    prob=model.predict_proba(np.array([[pct,wkt]]))[0]
                    rows.append({'Soal':soal,'Total':tot,'Benar':btr,'Persentase':f"{pct}%",
                                 'Waktu':int(wkt),'Prediksi':pred,
                                 f'P(Mudah)':f"{prob[0]*100:.1f}%",
                                 f'P(Sedang)':f"{prob[1]*100:.1f}%",
                                 f'P(Sulit)':f"{prob[2]*100:.1f}%"})
            if rows:
                st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
            else:
                st.warning("Format tidak dikenali. Gunakan: `NomorSoal, TotalSiswa, Benar, Waktu`")
        except Exception as e:
            st.error(f"Error: {e}")
    st.markdown('</div>', unsafe_allow_html=True)


def page_download():
    h = st.session_state.hasil
    st.markdown('<div class="card"><div class="card-title">💾 Download Hasil</div>', unsafe_allow_html=True)
    st.markdown("Unduh file hasil analisis dan model yang telah dilatih:")
    c1,c2=st.columns(2)
    with c1:
        st.markdown("### 📊 File Excel")
        st.markdown("Berisi semua sheet: Data, Iterasi K-Means, Grafik, Split, Training, Testing, Evaluasi")
        st.download_button("⬇️ Download Excel (.xlsx)",data=h['xl_bytes'],
                           file_name="hasil_kmeans_nb.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
    with c2:
        st.markdown("### 🤖 Model Pickle")
        st.markdown("Model Gaussian Naive Bayes yang telah dilatih, siap digunakan untuk prediksi")
        st.download_button("⬇️ Download Model (.pkl)",data=h['model_buf'],
                           file_name="model_nb.pkl",mime="application/octet-stream",
                           use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

    acc_pct=h['acc']*100
    st.markdown(f"""
    <div class="card">
    <div class="card-title">📈 Ringkasan Akhir</div>
    <div class="metric-row">
      <div class="metric-box"><div class="val">{len(h['data'])}</div><div class="lbl">Total Soal</div></div>
      <div class="metric-box green"><div class="val">{len(h['history'])}</div><div class="lbl">Iterasi K-Means</div></div>
      <div class="metric-box orange"><div class="val">{acc_pct:.1f}%</div><div class="lbl">Akurasi NB</div></div>
      <div class="metric-box red"><div class="val">{h['met_df']['F1'].mean():.3f}</div><div class="lbl">Macro F1</div></div>
    </div>
    </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN ROUTER
# ══════════════════════════════════════════════════════════════════════════════

def main():
    # Sidebar
    with st.sidebar:
        st.markdown("""
        <div style="text-align:center;padding:20px 0 10px;">
            <div style="font-size:2.5rem;">📊</div>
            <div style="font-size:1.1rem;font-weight:700;margin-top:6px;">Klasifikasi Soal</div>
            <div style="font-size:0.8rem;opacity:0.75;">K-Means + Naive Bayes</div>
        </div>
        <hr style="border-color:rgba(255,255,255,0.2);margin:12px 0;">
        """, unsafe_allow_html=True)
        
        has_data = st.session_state.hasil is not None
        
        pages_all = [
            ('upload',   '📤 Upload Data'),
            ('kmeans',   '🔵 Hasil K-Means'),
            ('model',    '🧠 Perbandingan Model'),
            ('evaluasi', '📈 Evaluasi Model'),
            ('prediksi', '🔮 Prediksi'),
            ('download', '💾 Download'),
        ]
        
        for key, label in pages_all:
            disabled = (not has_data and key not in ['upload'])
            if st.button(label, key=f"nav_{key}", use_container_width=True,
                         disabled=disabled):
                st.session_state.page = key
                st.rerun()
        
        if has_data:
            h=st.session_state.hasil
            st.markdown(f"""
            <hr style="border-color:rgba(255,255,255,0.2);margin:16px 0 10px;">
            <div style="font-size:0.8rem;opacity:0.8;padding:0 8px;line-height:2;">
            📌 Soal: <b>{len(h['data'])}</b><br>
            🔄 Iterasi: <b>{len(h['history'])}</b><br>
            🎯 Akurasi: <b>{h['acc']*100:.2f}%</b><br>
            🟢 Mudah: <b>{h['data']['Keterangan'].value_counts().get('Mudah',0)}</b><br>
            🟡 Sedang: <b>{h['data']['Keterangan'].value_counts().get('Sedang',0)}</b><br>
            🔴 Sulit: <b>{h['data']['Keterangan'].value_counts().get('Sulit',0)}</b>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div style="font-size:0.8rem;opacity:0.6;padding:12px 8px;text-align:center;">
            Upload file Excel terlebih dahulu untuk mengaktifkan menu lainnya.
            </div>""", unsafe_allow_html=True)

    # Router
    page = st.session_state.page
    if page == 'upload':   page_upload()
    elif not st.session_state.hasil:
        st.warning("⚠️ Silakan upload file Excel terlebih dahulu.")
        page_upload()
    elif page == 'kmeans':   page_kmeans()
    elif page == 'model':    page_model()
    elif page == 'evaluasi': page_evaluasi()
    elif page == 'prediksi': page_prediksi()
    elif page == 'download': page_download()

if __name__ == '__main__':
    main()